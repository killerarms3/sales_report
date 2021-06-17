import os
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Side, Border
from openpyxl.utils import get_column_letter

def get_driver(executable_path):
    options = Options()
    options.add_argument('-headless')
    driver = webdriver.Firefox(executable_path='/home/dytk2134/project/gitlab/sales_report/geckodriver', options=options)
    return driver

# 去CRM抓特定區間內的會員數
def get_members(tags, executable_path, username, password, start=None, end=None):
    driver = get_driver(executable_path)
    members_dict = {}
    try:
        driver.get('https://member.gsscloud.com/cas/login?service=https://www.videgree.com/23b0cab855374d7ebd4cb3eae90b0014/ExternalLogOn.mvc/SsoCasLoginCallback')
        driver.find_element_by_id('username').send_keys(username)
        driver.find_element_by_id('password').send_keys(password)
        driver.find_element_by_name('submit').click()

        for tag in tags:
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@id='LeftLabelQuickList']/div[@leftquicklabelid='Customer.All']"))).click()
            WebDriverWait(driver, 60).until(lambda driver: driver.execute_script("return jQuery.active == 0"))
            # time.sleep(5)
            # total
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='footer']/button[@class='btn more-btn']"))).click()
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='title']/button[@class='title-btn add']"))).click()
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='LabelItems']/label[text()='%s']" % (tag)))).click()
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'LabelConditionApply'))).click()
            WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, 'CreateDate'))).clear()
            if (start is not None) and (end is not None):
                WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, 'CreateDate'))).send_keys(start.strftime('%Y/%m/%d') + ' - ' + end.strftime('%Y/%m/%d'))
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='range_inputs']/button[@class='applyBtn hide btn btn-sm btn-success']"))).click()
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'QueryBtn'))).click()
            WebDriverWait(driver, 60).until(lambda driver: driver.execute_script("return jQuery.active == 0"))
            result_text =  WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div[@id='pager']/span[@class='k-pager-info k-label']"))).text
            try:
                total_member = int(re.search(r'共\s([0-9]+)\s筆', result_text).group(1))
            except AttributeError:
                total_member = 0
            if total_member != 0:
                # 消費次數大於0的會員
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='footer']/button[@class='btn more-btn']"))).click()
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.filter-group.order.hide-content'))).find_element_by_css_selector('button.filter-collapse-btn').click()
                WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, 'ConsumerTimes'))).send_keys('0')
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'QueryBtn'))).click()
                WebDriverWait(driver, 60).until(lambda driver: driver.execute_script("return jQuery.active == 0"))
                result_text =  WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div[@id='pager']/span[@class='k-pager-info k-label']"))).text
                try:
                    has_order = int(re.search(r'共\s([0-9]+)\s筆', result_text).group(1))
                except AttributeError:
                    has_order = 0
                if has_order != 0:
                    # 有消費但沒手機的會員
                    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='footer']/button[@class='btn more-btn']"))).click()
                    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.filter-group.customercontent.hide-content'))).find_element_by_css_selector('button.filter-collapse-btn').click()
                    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'select2-TelecomNumOption-container'))).click()
                    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//ul[@class='select2-results__options']/li[3]"))).click()
                    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='check-condition']/div[@class='checkbox-item']/label[@for='ReverseQueryTelecomNum']"))).click()
                    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'QueryBtn'))).click()
                    WebDriverWait(driver, 60).until(lambda driver: driver.execute_script("return jQuery.active == 0"))
                    result_text =  WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div[@id='pager']/span[@class='k-pager-info k-label']"))).text
                    try:
                        nophone_has_order = int(re.search(r'共\s([0-9]+)\s筆', result_text).group(1))
                    except AttributeError:
                        nophone_has_order = 0
            else:
                has_order = 0
                nophone_has_order = 0
            members_dict[tag] = {
                'total_member': total_member,
                'valid_member': has_order - nophone_has_order
            }

    finally:
        driver.quit()
    return members_dict

#將CRM查詢結果存至資料庫

# Group結果
def group_result(tags, members_dict):
    group = dict()
    for tag in tags:
        if tags[tag]['primary'] not in group:
            group[tags[tag]['primary']] = dict()
        if tags[tag]['secondary'] not in group[tags[tag]['primary']]:
            group[tags[tag]['primary']][tags[tag]['secondary']] = {
                'total_member': 0,
                'valid_member': 0
            }
        group[tags[tag]['primary']][tags[tag]['secondary']]['total_member'] += members_dict[tag]['total_member']
        group[tags[tag]['primary']][tags[tag]['secondary']]['valid_member'] += members_dict[tag]['valid_member']
    return group

def export_members(tags, template, output_file):
    # 之後改成由config抓資料
    executable_path='/home/dytk2134/project/gitlab/sales_report/geckodriver'
    username = 'IT@takkare.com'
    password = 'It52609575'
    # 取得會員數
    today = datetime.datetime.now()
    last_week_start = today - datetime.timedelta(days=today.weekday() + 7)
    last_week_end = today - datetime.timedelta(days=today.weekday() + 1)
    total_member = get_members(tags, executable_path, username, password)
    week_member = get_members(tags, executable_path, username, password, start=last_week_start, end=last_week_end)
    # 將week_member上傳至資料庫
    # 之後補上由資料庫查詢上上周與上上上週會員數

    # group
    group_total_member = group_result(tags, total_member)
    group_week_member = group_result(tags, week_member)

    # 讀取模板
    wb_template = openpyxl.load_workbook(template)
    member_sheet = wb_template['會員']
    valid_member_sheet = wb_template['有效會員']
    insert_row_count = 3 #由第三行前開始insert
    total_rows = [] # 小計行數
    gray_fill = PatternFill('solid', fgColor='D0CECE') # 灰色填滿
    bian = Side(style='thin', color='000000') # 設定邊框樣式
    border = Border(top=bian, bottom=bian, left=bian, right=bian)
    for primary in group_total_member:
        start_row = insert_row_count
        for secondary in group_total_member[primary]:
            # 插入空行
            member_sheet.insert_rows(insert_row_count)
            valid_member_sheet.insert_rows(insert_row_count)
            # 填入會員數
            member_sheet['A'+str(insert_row_count)] = secondary
            member_sheet['B'+str(insert_row_count)] = group_week_member[primary][secondary]['total_member']
            member_sheet['C'+str(insert_row_count)] = 0 # 上上週會員
            member_sheet['D'+str(insert_row_count)] = 0 # 上上上週會員
            member_sheet['E'+str(insert_row_count)] = group_total_member[primary][secondary]['total_member']
            valid_member_sheet['A'+str(insert_row_count)] = secondary
            valid_member_sheet['B'+str(insert_row_count)] = group_week_member[primary][secondary]['valid_member']
            valid_member_sheet['C'+str(insert_row_count)] = 0
            valid_member_sheet['D'+str(insert_row_count)] = 0
            valid_member_sheet['E'+str(insert_row_count)] = group_total_member[primary][secondary]['valid_member']
            insert_row_count += 1
        end_row = insert_row_count - 1
        for cell in sum(member_sheet['A'+str(start_row):'E'+str(end_row)], ()):
            cell.border = border
            valid_member_sheet[cell.coordinate].border = border
        member_sheet.insert_rows(insert_row_count)
        valid_member_sheet.insert_rows(insert_row_count)

        for letter in ['A', 'B', 'C', 'D', 'E']:
            if letter == 'A':
                member_sheet['A'+str(insert_row_count)] = primary + '小計'
                valid_member_sheet['A'+str(insert_row_count)] = primary + '小計'
            else:
                member_sheet[letter+str(insert_row_count)] = '= SUM(%s:%s)' % (letter+str(start_row), letter+str(end_row))
                valid_member_sheet[letter+str(insert_row_count)] = '= SUM(%s:%s)' % (letter+str(start_row), letter+str(end_row))
            member_sheet[letter+str(insert_row_count)].fill = gray_fill
            valid_member_sheet[letter+str(insert_row_count)].fill = gray_fill
            member_sheet[letter+str(insert_row_count)].border = border
            valid_member_sheet[letter+str(insert_row_count)].border = border
        total_rows.append(insert_row_count)
        insert_row_count += 1
    for letter in ['B', 'C', 'D', 'E']:
        member_sheet[letter+str(insert_row_count)] = '= SUM(%s)' % (', '.join([letter+str(c) for c in total_rows]))
        valid_member_sheet[letter+str(insert_row_count)] = '= SUM(%s)' % (', '.join([letter+str(c) for c in total_rows]))
    wb_template.save(output_file)

tags = dict()
with open('/home/dytk2134/project/gitlab/sales_report/tags.txt', 'r') as in_f:
    for line in in_f:
        if line and line[0] != '#':
            tokens = [t.strip() for t in line.split('\t')]
            tags[tokens[0]] = {
                'primary': tokens[1],
                'secondary': tokens[2]
            }


export_members(tags, '/home/dytk2134/project/gitlab/sales_report/templates/members.xlsx', '/home/dytk2134/project/gitlab/sales_report/members_result.xlsx')



