#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Side, Border
from decimal import Decimal

def invertory(excel_file, group, product):
    wb = openpyxl.load_workbook(excel_file)
    sheetname = wb.sheetnames[0]
    sheet = wb[sheetname]
    invertory_dict = dict()
    invertorybysku_dict = dict()
    for row in sheet.iter_rows(min_row=3):
        if row[0].value is not None:
            for idx, cell in enumerate(row[4:]):
                name = sheet[get_column_letter(idx+5)+'1'].value
                if name in group:
                    if group[name]['primary'] not in invertory_dict:
                        invertory_dict[group[name]['primary']] = dict()
                    if group[name]['secondary'] not in invertory_dict[group[name]['primary']]:
                        invertory_dict[group[name]['primary']][group[name]['secondary']] = Decimal(0.0)

                    if row[0].value not in invertorybysku_dict:
                        invertorybysku_dict[row[0].value] = {
                            'name': row[2].value,
                            'invertory': dict(),
                            'total': 0
                        }
                    if group[name]['secondary'] not in invertorybysku_dict[row[0].value]['invertory']:
                        invertorybysku_dict[row[0].value]['invertory'][group[name]['secondary']] = 0
                    invertorybysku_dict[row[0].value]['invertory'][group[name]['secondary']] += int(cell.value) if cell.value is not None else 0
                    invertorybysku_dict[row[0].value]['total'] += int(cell.value) if cell.value is not None else 0
                    # invertory_dict[group[name]['primary']][group[name]['secondary']]['goods'][row[0].value]['count'] += int(cell.value) if cell.value is not None else 0
                    invertory_dict[group[name]['primary']][group[name]['secondary']] += product[row[0].value] * int(cell.value) if cell.value is not None else 0
                else:
                    continue

    return invertory_dict, invertorybysku_dict

def export_invertory(invertory_dict, invertorybysku_dict, template, output_file):
    # 讀取模板
    wb_template = openpyxl.load_workbook(template)
    invertory_sheet = wb_template['Invertory']
    invertorybysku_sheet = wb_template['Inventory by SKU']
    insert_row_count = 3
    total_rows = [] # 小計行數
    gray_fill = PatternFill('solid', fgColor='D0CECE') # 灰色填滿
    bian = Side(style='thin', color='000000') # 設定邊框樣式
    border = Border(top=bian, bottom=bian, left=bian, right=bian)
    secondary_invertory = list()
    # 輸出Invertory
    for primary in invertory_dict:
        start_row = insert_row_count
        for secondary in invertory_dict[primary]:
            secondary_invertory.append(secondary)
            invertory_sheet.insert_rows(insert_row_count)
            invertory_sheet['A'+str(insert_row_count)] = secondary
            invertory_sheet['B'+str(insert_row_count)] = invertory_dict[primary][secondary]
            invertory_sheet['C'+str(insert_row_count)] = invertory_dict[primary][secondary] - 0 #對比上週(本週-上週)
            insert_row_count += 1
        end_row = insert_row_count - 1
        for cell in sum(invertory_sheet['A'+str(start_row):'C'+str(end_row)], ()):
            cell.border = border
        invertory_sheet.insert_rows(insert_row_count)
        for letter in ['A', 'B', 'C']:
            if letter == 'A':
                invertory_sheet['A'+str(insert_row_count)] = primary + '小計'
            else:
                invertory_sheet[letter+str(insert_row_count)] = '= SUM(%s:%s)' % (letter+str(start_row), letter+str(end_row))
            invertory_sheet[letter+str(insert_row_count)].fill = gray_fill
            invertory_sheet[letter+str(insert_row_count)].border = border
        total_rows.append(insert_row_count)
        insert_row_count += 1
    for letter in ['B', 'C']:
        invertory_sheet[letter+str(insert_row_count)] = '= SUM(%s)' % (', '.join([letter+str(c) for c in total_rows]))
    # 輸出Invertory By SKU
    insert_row_count = 3
    for idx, sku in enumerate(sorted(invertorybysku_dict.keys(), key=lambda k: invertorybysku_dict[k]['total'], reverse=True)):
        if any(invertorybysku_dict[sku]['invertory'].values()):
            invertorybysku_sheet['A'+str(insert_row_count)] = sku
            invertorybysku_sheet['A'+str(insert_row_count)].border = border
            invertorybysku_sheet['B'+str(insert_row_count)] = invertorybysku_dict[sku]['name']
            invertorybysku_sheet['B'+str(insert_row_count)].border = border
            for idy, invertory in enumerate(secondary_invertory):
                invertorybysku_sheet[get_column_letter(4+idy)+'2'] = invertory
                invertorybysku_sheet[get_column_letter(4+idy)+'2'].border = border
                invertorybysku_sheet[get_column_letter(4+idy)+str(insert_row_count)] = invertorybysku_dict[sku]['invertory'][invertory]
                invertorybysku_sheet[get_column_letter(4+idy)+str(insert_row_count)].border = border
            insert_row_count += 1

    wb_template.save(output_file)

group_dict = dict()
product_dict = dict()

with open('/home/dytk2134/project/gitlab/sales_report/stores.txt', 'r') as in_f:
    for line in in_f:
        if line and line[0] != '#':
            tokens = [t.strip() for t in line.split('\t')]
            group_dict[tokens[1]] = {
                'primary': tokens[2],
                'secondary': tokens[3]
            }

with open('/home/dytk2134/project/gitlab/sales_report/items.txt', 'r') as in_f:
    for line in in_f:
        if line and line[0] != '#':
            tokens = [t.strip() for t in line.split('\t')]
            product_dict[tokens[0]] = Decimal(tokens[2])

invertory_dict, invertorybysku_dict = invertory('/home/dytk2134/project/gitlab/sales_report/現有庫存量查詢作業.xlsx', group_dict, product_dict)
export_invertory(invertory_dict, invertorybysku_dict,'/home/dytk2134/project/gitlab/sales_report/templates/invertory.xlsx', 'invertory_output.xlsx')