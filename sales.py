import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Side, Border
import datetime
from decimal import Decimal
import string
import sys

def daily_sales(sales_excel, prepaid_excel, prepaid_notship_excel, stores_dict):
    # stores_dict = {
    #     '店櫃名稱': '群組'
    #     '內湖專櫃': '內湖',
    #     '經銷商-HOLA': '經銷-特力(H+C+P)'
    # }

    date_format = "%Y/%m/%d"
    # 會員日銷費明細表
    sales_wb = openpyxl.load_workbook(sales_excel)
    sales_sheet = sales_wb[sales_wb.sheetnames[0]]
    daily_sales_dict = dict()
    for row in sales_sheet.values:
        if row[0] == '會員代號':
            continue
        else:
            # 目前全部都計算，之後須確認是否需要忽略某些客戶的購買紀錄
            try:
                datetime.datetime.strptime(row[0], date_format)
                date = row[0].replace('/', '')
                if date not in daily_sales_dict:
                    daily_sales_dict[date] = dict()
                # 商品代號
                if row[7] not in daily_sales_dict[date]:
                    daily_sales_dict[date][row[7]] = {
                        'name': row[8], # 產品名稱
                        'total': 0, # 小計
                        'store': dict()
                    }
                # 如果不在名單，都列在'未分類'
                group = stores_dict[row[3]] if row[3] in stores_dict else '未分類'
                if group not in daily_sales_dict[date][row[7]]['store']:
                    daily_sales_dict[date][row[7]]['store'][group] = 0
                daily_sales_dict[date][row[7]]['total'] += int(row[12])
                daily_sales_dict[date][row[7]]['store'][group] += int(row[12]) # 訂購數量
            except ValueError:
                continue
    # 訂金查詢作業
    prepaid_wb = openpyxl.load_workbook(prepaid_excel)
    prepaid_sheet = prepaid_wb[prepaid_wb.sheetnames[0]]
    # 訂金未出貨商品明細表
    prepaid_notship_wb = openpyxl.load_workbook(prepaid_notship_excel)
    prepaid_notship_sheet = prepaid_notship_wb[prepaid_notship_wb.sheetnames[0]]
    # 抓還未銷貨的紀錄，若已銷貨，此資料會出現在會員日消費明細表
    prepaid_dict = dict()
    for row in prepaid_sheet.values:
        try:
            if row[2] is not None:
                datetime.datetime.strptime(row[2], date_format)
                # 沒有銷貨編號
                if row[12] is None:
                    # (row[1], row[3], row[4]): (店櫃名稱, 會員卡號, 銷貨編號(訂金))
                    prepaid_dict[(row[1], row[3], row[4])] =  row[2].replace('/', '') # 銷貨日期(訂金)
        except ValueError:
            continue
    for row in prepaid_notship_sheet.values:
        # 假設銷貨編號全都是數字
        try:
            if row[3] is not None:
                int(row[3])
                store = ' '.join(row[2].split(' ')[1:])
                membercard = ' '.join(row[4].split(' ')[1:])
                if (store, membercard, row[3]) in prepaid_dict:
                    date = prepaid_dict[(store, membercard, row[3])]
                    if date not in daily_sales_dict:
                        daily_sales_dict[date] = dict()
                    if row[0] not in daily_sales_dict[date]:
                        daily_sales_dict[date][row[0]] = {
                            'name': row[1], # 產品名稱
                            'total': 0, # 小計
                            'store': dict()
                        }
                    group = stores_dict[store] if store in stores_dict else '其他'
                    if group not in daily_sales_dict[date][row[0]]['store']:
                        daily_sales_dict[date][row[0]]['store'][group] = 0
                    daily_sales_dict[date][row[0]]['store'][group] += int(row[7]) # 訂購數量
            else:
                continue
        except ValueError:
            continue
    return daily_sales_dict

def export_daily_sales(group_list, daily_sales_dict, template, output_postfix):
    # style
    bian = Side(style='thin', color='000000') # 設定邊框樣式
    border = Border(top=bian, bottom=bian, left=bian, right=bian)
    # 輸出日銷售清單
    for date in daily_sales_dict:
        insert_row_count = 4
        # 讀取模板
        wb_template = openpyxl.load_workbook(template)
        daily_sales_sheet = wb_template['日銷售清單']
        output_file = date + output_postfix
        for idx, sku in enumerate(sorted(daily_sales_dict[date].keys(), key=lambda k: daily_sales_dict[date][k]['total'], reverse=True)):
            daily_sales_sheet['A'+str(insert_row_count)] = sku
            daily_sales_sheet['A'+str(insert_row_count)].border = border
            daily_sales_sheet['B'+str(insert_row_count)] = daily_sales_dict[date][sku]['name']
            daily_sales_sheet['B'+str(insert_row_count)].border = border
            for idy, group in enumerate(group_list):
                daily_sales_sheet[get_column_letter(4+idy)+'3'] = group
                daily_sales_sheet[get_column_letter(4+idy)+'3'].border = border
                daily_sales_sheet[get_column_letter(4+idy)+str(insert_row_count)] = daily_sales_dict[date][sku]['store'][group] if group in daily_sales_dict[date][sku]['store'] else 0
                daily_sales_sheet[get_column_letter(4+idy)+str(insert_row_count)].border = border
            daily_sales_sheet[get_column_letter(4+len(group_list))+'3'] = '小計'
            daily_sales_sheet[get_column_letter(4+len(group_list))+'3'].border = border
            daily_sales_sheet[get_column_letter(4+len(group_list))+str(insert_row_count)] = daily_sales_dict[date][sku]['total']
            daily_sales_sheet[get_column_letter(4+len(group_list))+str(insert_row_count)].border = border
            insert_row_count += 1
        wb_template.save(output_file)

def sales(product_dict, sales_excel, prepaid_excel, prepaid_notship_excel):
    date_format = "%Y/%m/%d"
    # 會員日銷費明細表
    sales_wb = openpyxl.load_workbook(sales_excel)
    sales_sheet = sales_wb[sales_wb.sheetnames[0]]
    sales_dict = dict()
    for row in sales_sheet.values:
        if row[0] == '會員代號':
            continue
        else:
            # 目前全部都計算，之後須確認是否需要忽略某些客戶的購買紀錄
            try:
                date = datetime.datetime.strptime(row[0], date_format)
                if date not in sales_dict:
                    sales_dict[date] = dict()
                # 消費店櫃
                store = row[3].strip()
                if store not in sales_dict[date]:
                    sales_dict[date][store] = {
                        'tickets': set(),
                        'sales': 0,
                        'cost': 0
                    }
                sales_dict[date][store]['tickets'].add((row[0], store, row[1], row[6])) # 日期, 消費店櫃, 會員卡號, 銷貨編號
                sales_dict[date][store]['sales'] += Decimal(row[13]) # 金額
                if row[7] not in product_dict:
                    print('無法找到商品編號為%s的商品資訊，請更新商品表' % (row[7]))
                    sys.exit(1)
                else:
                    sales_dict[date][store]['cost'] += product_dict[row[7]] * int(row[12]) # 成本 * 數量
            except ValueError:
                continue
    # 訂金查詢作業
    prepaid_wb = openpyxl.load_workbook(prepaid_excel)
    prepaid_sheet = prepaid_wb[prepaid_wb.sheetnames[0]]
    # 訂金未出貨商品明細表
    prepaid_notship_wb = openpyxl.load_workbook(prepaid_notship_excel)
    prepaid_notship_sheet = prepaid_notship_wb[prepaid_notship_wb.sheetnames[0]]
    # 抓還未銷貨的紀錄，若已銷貨，此資料會出現在會員日消費明細表
    prepaid_dict = dict()
    for row in prepaid_sheet.values:
        try:
            if row[2] is not None:
                datetime.datetime.strptime(row[2], date_format)
                # 沒有銷貨編號
                if row[12] is None:
                    # (row[1], row[3], row[4]): (店櫃名稱, 會員卡號, 銷貨編號(訂金))
                    prepaid_dict[(row[1], row[3], row[4])] = {
                        'date': row[2].replace('/', ''), # 銷貨日期(訂金)
                        'price': int(row[6]) + int(row[7]) # 訂金金額+尾款金額
                    }
        except ValueError:
            continue
    for row in prepaid_notship_sheet.values:
        # 假設銷貨編號全都是數字
        try:
            if row[3] is not None:
                int(row[3])
                store = ' '.join(row[2].split(' ')[1:]).strip()
                membercard = ' '.join(row[4].split(' ')[1:])
                if (store, membercard, row[3]) in prepaid_dict:
                    date = prepaid_dict[(store, membercard, row[3])]
                    if date not in sales_dict:
                        sales_dict[date] = dict()
                    if store not in sales_dict[date]:
                        sales_dict[date][store] = {
                            'tickets': set(),
                            'sales': 0,
                            'cost': 0
                        }
                    sales_dict[date][store]['tickets'].add((prepaid_dict[(store, membercard, row[3])]['date'], store, membercard, row[3])) # 日期, 消費店櫃, 會員卡號, 銷貨編號
                    sales_dict[date][store]['sales'] += prepaid_dict[(store, membercard, row[3])]['price'] # 金額
                    sales_dict[date][store]['cost'] += product_dict[row[0]] * int(row[7]) # 成本 * 數量
        except ValueError:
            continue
    return sales_dict

def group_sales(start_date, end_date, group_dict, sales_dict):
    result_dict = dict()
    for date in sales_dict:
        if date >= start_date and date <= end_date:
            for store in sales_dict[date]:
                store = store.strip()
                if group_dict[store]['primary'] not in result_dict:
                    result_dict[group_dict[store]['primary']] = dict()
                if store not in result_dict[group_dict[store]['primary']]:
                    result_dict[group_dict[store]['primary']][store] = {
                        'tickets_num': 0,
                        'sales': 0,
                        'cost': 0
                    }
                result_dict[group_dict[store]['primary']][store]['tickets_num'] += len(sales_dict[date][store]['tickets'])
                result_dict[group_dict[store]['primary']][store]['sales'] += sales_dict[date][store]['sales']
                result_dict[group_dict[store]['primary']][store]['cost'] += sales_dict[date][store]['cost']
    return result_dict


def export_sales(sales_budget, margin_budget, result_dict, template, output_file):
    # style
    gray_fill = PatternFill('solid', fgColor='D0CECE') # 灰色填滿
    bian = Side(style='thin', color='000000') # 設定邊框樣式
    border = Border(top=bian, bottom=bian, left=bian, right=bian)
    # 讀取模板
    wb_template = openpyxl.load_workbook(template)
    sales_sheet = wb_template[wb_template.sheetnames[0]]
    insert_row_count = 3
    total_rows = [] # 小計行數
    for primary in result_dict:
        start_row = insert_row_count
        for secondary in result_dict[primary]:
            sales_sheet.insert_rows(insert_row_count)
            sales_sheet['A'+str(insert_row_count)] = secondary # Store
            sales_sheet['B'+str(insert_row_count)] = result_dict[primary][secondary]['sales'] # Sales $
            sales_sheet['C'+str(insert_row_count)] = sales_budget # Sales $ Budget
            sales_sheet['D'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('B'+str(insert_row_count), 'C'+str(insert_row_count)) # Sales$ ach%
            sales_sheet['E'+str(insert_row_count)] = 100 # Sales$ LFL%
            sales_sheet['F'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('G'+str(insert_row_count), 'B'+str(insert_row_count)) # Margin%
            sales_sheet['G'+str(insert_row_count)] = result_dict[primary][secondary]['sales'] - result_dict[primary][secondary]['cost'] # Margin$
            sales_sheet['H'+str(insert_row_count)] = margin_budget # Margin$ Budget
            sales_sheet['I'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('G'+str(insert_row_count), 'H'+str(insert_row_count)) # Margin$ ach%
            sales_sheet['J'+str(insert_row_count)] = 100 # Margin$ LFL%
            sales_sheet['K'+str(insert_row_count)] = result_dict[primary][secondary]['tickets_num'] # Tickets #
            sales_sheet['L'+str(insert_row_count)] = 100 # Tickets# LFL%
            insert_row_count += 1
        end_row = insert_row_count - 1
        for cell in sum(sales_sheet['A'+str(start_row):'L'+str(end_row)], ()):
            cell.border = border
        sales_sheet.insert_rows(insert_row_count)
        sales_sheet['A'+str(insert_row_count)] = primary + '小計'
        sales_sheet['D'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('B'+str(insert_row_count), 'C'+str(insert_row_count))
        sales_sheet['F'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('G'+str(insert_row_count), 'B'+str(insert_row_count))
        sales_sheet['I'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('G'+str(insert_row_count), 'H'+str(insert_row_count))
        for letter in list(string.ascii_uppercase)[0:12]:
            if letter in {'B', 'C', 'G', 'H', 'K'}:
                sales_sheet[letter+str(insert_row_count)] = '= SUM(%s:%s)' % (letter+str(start_row), letter+str(end_row))
            sales_sheet[letter+str(insert_row_count)].fill = gray_fill
            sales_sheet[letter+str(insert_row_count)].border = border
        total_rows.append(insert_row_count)
        insert_row_count += 1
    # 總計
    sales_sheet['D'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('B'+str(insert_row_count), 'C'+str(insert_row_count))
    sales_sheet['F'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('G'+str(insert_row_count), 'B'+str(insert_row_count))
    sales_sheet['I'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('G'+str(insert_row_count), 'H'+str(insert_row_count))
    for letter in ['B', 'C', 'G', 'H', 'K']:
        sales_sheet[letter+str(insert_row_count)] = '= SUM(%s)' % (', '.join([letter+str(c) for c in total_rows]))
    wb_template.save(output_file)



group_dict = dict()
group_list = list()
product_dict = dict()

with open('/home/dytk2134/project/gitlab/sales_report/stores.txt', 'r') as in_f:
    for line in in_f:
        if line and line[0] != '#':
            tokens = [t.strip() for t in line.split('\t')]
            group_dict[tokens[1]] = tokens[3]
            if tokens[3] not in group_list:
                group_list.append(tokens[3])
group_list.append('其他')

daily_sales_dict = daily_sales('會員日期商品消費明細表.xlsx', '訂金查詢作業.xlsx', '訂金未出貨商品明細表.xlsx', group_dict)
export_daily_sales(group_list, daily_sales_dict, 'templates/daily_sales.xlsx', '日銷售清單.xlsx')
with open('/home/dytk2134/project/gitlab/sales_report/items.txt', 'r') as in_f:
    for line in in_f:
        if line and line[0] != '#':
            tokens = [t.strip() for t in line.split('\t')]
            product_dict[tokens[0]] = Decimal(tokens[2])

group_dict = dict()
with open('/home/dytk2134/project/gitlab/sales_report/stores.txt', 'r') as in_f:
    for line in in_f:
        if line and line[0] != '#':
            tokens = [t.strip() for t in line.split('\t')]
            group_dict[tokens[1]] = {
                'primary': tokens[2],
                'secondary': tokens[3]
            }


sales_dict = sales(product_dict, '會員日期商品消費明細表.xlsx', '訂金查詢作業.xlsx', '訂金未出貨商品明細表.xlsx')
start_date = datetime.datetime(2021, 5, 5)
end_date = datetime.datetime(2021, 5, 16)

result_dict = group_sales(start_date, end_date, group_dict, sales_dict)
export_sales(10000, 10000, result_dict, 'templates/sales.xlsx', '202010505_20210516銷售總表.xlsx')