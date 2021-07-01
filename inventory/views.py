from django.shortcuts import render, redirect
from django.urls import reverse
from django.http import HttpResponse, Http404
from django.views.decorators.csrf import csrf_protect
from inventory.models import Inventory
from SKU_report.models import DailyInventoryBySKU
from extra_table.models import Store_house, SKU
from django.conf import settings
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Side, Border
from decimal import Decimal
import datetime
import os
import glob
import re
# Create your views here.

def save_inventorybysku(excel_file):
    today = datetime.datetime.now()
    last_1_week_date = today - datetime.timedelta(days=today.weekday() + 1) + datetime.timedelta(days=1)
    wb = openpyxl.load_workbook(excel_file)
    sheetname = wb.sheetnames[0]
    sheet = wb[sheetname]
    for row in sheet.iter_rows(min_row=3):
        if row[0].value is not None:
            for idx, cell in enumerate(row[4:]):
                name = sheet[get_column_letter(idx+5)+'1'].value
                if Store_house.objects.filter(name=name) and SKU.objects.filter(sku=row[0].value):
                    store_house = Store_house.objects.filter(name=name)[0]
                    sku = SKU.objects.get(sku=row[0].value)
                    if DailyInventoryBySKU.objects.filter(date=last_1_week_date, store_house=store_house, sku=sku):
                        dailyinventorybysku = DailyInventoryBySKU.objects.get(date=last_1_week_date, store_house=store_house, sku=sku)
                    else:
                        dailyinventorybysku = DailyInventoryBySKU(date=last_1_week_date, store_house=store_house, sku=sku)
                    if cell.value is not None:
                        dailyinventorybysku.counts = int(cell.value)
                        dailyinventorybysku.save()
                else:
                    continue

def generate_inventory(last_1_week_date):
    inventory_dict = dict()
    for dailyinventorybysku in DailyInventoryBySKU.objects.filter(date=last_1_week_date):
        if dailyinventorybysku.store_house not in inventory_dict:
            inventory_dict[dailyinventorybysku.store_house] = 0
        inventory_dict[dailyinventorybysku.store_house] += dailyinventorybysku.sku.cost
    # save inventory
    for store_house in inventory_dict:
        if not Inventory.objects.filter(date=last_1_week_date, store_house=store_house):
            inventory = Inventory(date=last_1_week_date, store_house=store_house)
        else:
            inventory = Inventory.objects.get(date=last_1_week_date, store_house=store_house)
        inventory.inventory = inventory_dict[store_house]
        inventory.save()

def group_inventorybysku(date):
    result_dict = dict()
    subtypes = set()
    for store_house in Store_house.objects.all():
        for inventorybysku in DailyInventoryBySKU.objects.filter(date=date, store_house=store_house):
            subtypes.add(inventorybysku.store_house.subtype)
            if inventorybysku.sku not in result_dict:
                result_dict[inventorybysku.sku] = {
                    'inventory': dict(),
                    'total': 0
                }
            if inventorybysku.store_house.subtype not in result_dict:
                result_dict[inventorybysku.sku]['inventory'][inventorybysku.store_house.subtype] = 0
            result_dict[inventorybysku.sku]['inventory'][inventorybysku.store_house.subtype] += inventorybysku.counts
            result_dict[inventorybysku.sku]['total'] += inventorybysku.counts
    return result_dict, list(subtypes)


def group_inventory(date):
    result_dict = dict()
    for store_house in Store_house.objects.all():
        inventorys = Inventory.objects.filter(date=date.strftime('%Y-%m-%d'), store_house=store_house)
        if store_house.category not in result_dict:
            result_dict[store_house.category] = dict()
        if store_house.subtype not in result_dict[store_house.category]:
            result_dict[store_house.category][store_house.subtype] = 0
        if inventorys:
            inventory = inventorys[0]
            result_dict[store_house.category][store_house.subtype] += inventory.inventory
    return result_dict

@csrf_protect
def upload_inventory(request):
    errors = list()
    if request.method == 'POST':
        upload_file = request.FILES['inventory']
        save_inventorybysku(upload_file)
        today = datetime.datetime.now()
        last_1_week_date = today - datetime.timedelta(days=today.weekday() + 1) + datetime.timedelta(days=1)
        generate_inventory(last_1_week_date)
        export_inventory(last_1_week_date)
        return redirect(reverse('inventory:inventory_report'))
    return render(request,'upload_inventory.html', locals())

def export_inventory(last_1_week_date):
    template = os.path.join(settings.BASE_DIR, 'excel_templates', 'inventory.xlsx')
    output_file = os.path.join(settings.BASE_DIR, 'output', 'Inventory%s.xlsx' % last_1_week_date.strftime('%Y%m%d'))
    # 讀取模板
    wb_template = openpyxl.load_workbook(template)
    inventory_sheet = wb_template['Inventory']
    inventorybysku_sheet = wb_template['Inventory by SKU']
    insert_row_count = 3
    total_rows = [] # 小計行數
    gray_fill = PatternFill('solid', fgColor='D0CECE') # 灰色填滿
    bian = Side(style='thin', color='000000') # 設定邊框樣式
    border = Border(top=bian, bottom=bian, left=bian, right=bian)
    # 輸出Inventory
    last_2_week_date = last_1_week_date - datetime.timedelta(days=7)
    last_1_week_result = group_inventory(last_1_week_date)
    last_2_week_result = group_inventory(last_2_week_date)

    for primary in last_1_week_result:
        # 上週資料
        start_row = insert_row_count
        for secondary in last_1_week_result[primary]:
            inventory_sheet.insert_rows(insert_row_count)
            inventory_sheet['A'+str(insert_row_count)] = secondary
            inventory_sheet['B'+str(insert_row_count)] = last_1_week_result[primary][secondary]
            inventory_sheet['C'+str(insert_row_count)] = last_1_week_result[primary][secondary] - last_2_week_result[primary][secondary] #對比上週(本週-上週)
            insert_row_count += 1
        end_row = insert_row_count - 1
        for cell in sum(inventory_sheet['A'+str(start_row):'C'+str(end_row)], ()):
            cell.border = border
        inventory_sheet.insert_rows(insert_row_count)
        for letter in ['A', 'B', 'C']:
            if letter == 'A':
                inventory_sheet['A'+str(insert_row_count)] = primary + '小計'
            else:
                inventory_sheet[letter+str(insert_row_count)] = '= SUM(%s:%s)' % (letter+str(start_row), letter+str(end_row))
            inventory_sheet[letter+str(insert_row_count)].fill = gray_fill
            inventory_sheet[letter+str(insert_row_count)].border = border
        total_rows.append(insert_row_count)
        insert_row_count += 1
    for letter in ['B', 'C']:
        inventory_sheet[letter+str(insert_row_count)] = '= SUM(%s)' % (', '.join([letter+str(c) for c in total_rows]))

    # 輸出Inventory By SKU
    insert_row_count = 3
    inventorybysku_dict, subtype_list = group_inventorybysku(last_1_week_date)
    for sku in sorted(inventorybysku_dict.keys(), key=lambda k: inventorybysku_dict[k]['total'], reverse=True):
        if any(inventorybysku_dict[sku]['inventory'].values()):
            status = ''
            inventorybysku_sheet['A'+str(insert_row_count)] = sku.sku
            inventorybysku_sheet['A'+str(insert_row_count)].border = border
            inventorybysku_sheet['B'+str(insert_row_count)] = sku.name
            inventorybysku_sheet['B'+str(insert_row_count)].border = border
            inventorybysku_sheet['C'+str(insert_row_count)] = sku.status
            inventorybysku_sheet['C'+str(insert_row_count)].border = border
            for idy, subtype in enumerate(subtype_list):
                inventorybysku_sheet[get_column_letter(4+idy)+'2'] = subtype
                inventorybysku_sheet[get_column_letter(4+idy)+'2'].border = border
                inventorybysku_sheet[get_column_letter(4+idy)+str(insert_row_count)] = inventorybysku_dict[sku]['inventory'][subtype] if subtype in inventorybysku_dict[sku]['inventory'] else 0
                inventorybysku_sheet[get_column_letter(4+idy)+str(insert_row_count)].border = border
            insert_row_count += 1
    wb_template.save(output_file)

@csrf_protect
def inventory_report(request):
    reports = dict()
    if request.method == 'POST':
        date = datetime.datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
        last_1_week_date = date - datetime.timedelta(days=date.weekday() + 1) + datetime.timedelta(days=1)
        generate_inventory(last_1_week_date)
        export_inventory(last_1_week_date)
        return redirect(reverse('inventory:inventory_report'))
    else:
        for file_path in glob.glob(os.path.join(settings.BASE_DIR, 'output', 'Inventory*.xlsx')):
            filename = os.path.basename(file_path)
            date = re.findall(r'\d+', filename)[0]
            reports[date] = filename
    return render(request,'inventory_report.html', locals())

def download(request, filename):
    file_path = os.path.join(settings.BASE_DIR, 'output', filename)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404
