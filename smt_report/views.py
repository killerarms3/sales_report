from django.shortcuts import render, redirect
from django.urls import reverse
from django.http import HttpResponse, Http404
from django.views.decorators.csrf import csrf_protect
from django.conf import settings
from smt_report.models import smt_report
from SKU_report.models import DailyInventoryBySKU
from extra_table.models import Store_house, SKU, Stores, Stores_budget
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Side, Border
from decimal import Decimal
import string
import datetime
import os
import glob
import re
from utils.utils import weektodate, monthly, yeartodate, lineforline
from django.apps import apps

# Create your views here.

def daily_sales(sales_excel, prepaid_excel, prepaid_notship_excel, stores_dict):
    # stores_dict = {
    #     '店櫃名稱': '群組'
    #     '內湖專櫃': '內湖',
    #     '經銷商-HOLA': '經銷-特力(H+C+P)'
    # }

    date_format = "%Y/%m/%d"
    # 會員日銷費明細表
    sales_wb = openpyxl.load_workbook(sales_excel)
    sales_sheet = sales_wb[sales_wb.sheetnames[0]]
    daily_sales_dict = dict()
    for row in sales_sheet.values:
        if row[0] == '會員代號':
            continue
        else:
            # 目前全部都計算，之後須確認是否需要忽略某些客戶的購買紀錄
            try:
                datetime.datetime.strptime(row[0], date_format)
                date = row[0].replace('/', '')
                if date not in daily_sales_dict:
                    daily_sales_dict[date] = dict()
                # 商品代號
                if row[7] not in daily_sales_dict[date]:
                    daily_sales_dict[date][row[7]] = {
                        'name': row[8], # 產品名稱
                        'total': 0, # 小計
                        'store': dict()
                    }
                # 如果不在名單，都列在'未分類'
                group = stores_dict[row[3]] if row[3] in stores_dict else '未分類'
                if group not in daily_sales_dict[date][row[7]]['store']:
                    daily_sales_dict[date][row[7]]['store'][group] = 0
                daily_sales_dict[date][row[7]]['total'] += int(row[12])
                daily_sales_dict[date][row[7]]['store'][group] += int(row[12]) # 訂購數量
            except ValueError:
                continue
    # 訂金查詢作業
    prepaid_wb = openpyxl.load_workbook(prepaid_excel)
    prepaid_sheet = prepaid_wb[prepaid_wb.sheetnames[0]]
    # 訂金未出貨商品明細表
    prepaid_notship_wb = openpyxl.load_workbook(prepaid_notship_excel)
    prepaid_notship_sheet = prepaid_notship_wb[prepaid_notship_wb.sheetnames[0]]
    # 抓還未銷貨的紀錄，若已銷貨，此資料會出現在會員日消費明細表
    prepaid_dict = dict()
    for row in prepaid_sheet.values:
        try:
            if row[2] is not None:
                datetime.datetime.strptime(row[2], date_format)
                # 沒有銷貨編號
                if row[12] is None:
                    # (row[1], row[3], row[4]): (店櫃名稱, 會員卡號, 銷貨編號(訂金))
                    prepaid_dict[(row[1], row[3], row[4])] =  row[2].replace('/', '') # 銷貨日期(訂金)
        except ValueError:
            continue
    for row in prepaid_notship_sheet.values:
        # 假設銷貨編號全都是數字
        try:
            if row[3] is not None:
                int(row[3])
                store = ' '.join(row[2].split(' ')[1:])
                membercard = ' '.join(row[4].split(' ')[0])
                if (store, membercard, row[3]) in prepaid_dict:
                    date = prepaid_dict[(store, membercard, row[3])]
                    if date not in daily_sales_dict:
                        daily_sales_dict[date] = dict()
                    if row[0] not in daily_sales_dict[date]:
                        daily_sales_dict[date][row[0]] = {
                            'name': row[1], # 產品名稱
                            'total': 0, # 小計
                            'store': dict()
                        }
                    group = stores_dict[store] if store in stores_dict else '其他'
                    if group not in daily_sales_dict[date][row[0]]['store']:
                        daily_sales_dict[date][row[0]]['store'][group] = 0
                    daily_sales_dict[date][row[0]]['store'][group] += int(row[7]) # 訂購數量
            else:
                continue
        except ValueError:
            continue
    return daily_sales_dict

def export_daily_sales(group_list, daily_sales_dict, template, output_postfix):
    # style
    bian = Side(style='thin', color='000000') # 設定邊框樣式
    border = Border(top=bian, bottom=bian, left=bian, right=bian)
    # 輸出日銷售清單
    for date in daily_sales_dict:
        insert_row_count = 4
        # 讀取模板
        wb_template = openpyxl.load_workbook(template)
        daily_sales_sheet = wb_template['日銷售清單']
        output_file =  output_postfix
        for idx, sku in enumerate(sorted(daily_sales_dict[date].keys(), key=lambda k: daily_sales_dict[date][k]['total'], reverse=True)):
            daily_sales_sheet['A'+str(insert_row_count)] = sku
            daily_sales_sheet['A'+str(insert_row_count)].border = border
            daily_sales_sheet['B'+str(insert_row_count)] = daily_sales_dict[date][sku]['name']
            daily_sales_sheet['B'+str(insert_row_count)].border = border
            for idy, group in enumerate(group_list):
                daily_sales_sheet[get_column_letter(4+idy)+'3'] = group
                daily_sales_sheet[get_column_letter(4+idy)+'3'].border = border
                daily_sales_sheet[get_column_letter(4+idy)+str(insert_row_count)] = daily_sales_dict[date][sku]['store'][group] if group in daily_sales_dict[date][sku]['store'] else 0
                daily_sales_sheet[get_column_letter(4+idy)+str(insert_row_count)].border = border
            daily_sales_sheet[get_column_letter(4+len(group_list))+'3'] = '小計'
            daily_sales_sheet[get_column_letter(4+len(group_list))+'3'].border = border
            daily_sales_sheet[get_column_letter(4+len(group_list))+str(insert_row_count)] = daily_sales_dict[date][sku]['total']
            daily_sales_sheet[get_column_letter(4+len(group_list))+str(insert_row_count)].border = border
            insert_row_count += 1
        wb_template.save(output_file)

def sales(product_dict, sales_excel, prepaid_excel, prepaid_notship_excel):
    date_format = "%Y/%m/%d"
    # 會員日銷費明細表
    sales_wb = openpyxl.load_workbook(sales_excel)
    sales_sheet = sales_wb[sales_wb.sheetnames[0]]
    sales_dict = dict()
    for row in sales_sheet.values:
        if row[0] == '會員代號':
            continue
        else:
            # 目前全部都計算，之後須確認是否需要忽略某些客戶的購買紀錄
            try:
                date = datetime.datetime.strptime(row[0], date_format)
                if date not in sales_dict:
                    sales_dict[date] = dict()
                # 消費店櫃
                store = row[3].strip()
                if store not in sales_dict[date]:
                    sales_dict[date][store] = {
                        'tickets': set(),
                        'sales': 0,
                        'cost': 0
                    }
                sales_dict[date][store]['tickets'].add((row[0], store, row[1], row[6])) # 日期, 消費店櫃, 會員卡號, 銷貨編號
                sales_dict[date][store]['sales'] += Decimal(row[13]) # 金額
                if row[7] not in product_dict:
                    print('無法找到商品編號為%s的商品資訊，請更新商品表' % (row[7]))
                    sys.exit(1)
                else:
                    sales_dict[date][store]['cost'] += product_dict[row[7]] * int(row[12]) # 成本 * 數量
            except ValueError:
                continue
    # 訂金查詢作業
    prepaid_wb = openpyxl.load_workbook(prepaid_excel)
    prepaid_sheet = prepaid_wb[prepaid_wb.sheetnames[0]]
    # 訂金未出貨商品明細表
    prepaid_notship_wb = openpyxl.load_workbook(prepaid_notship_excel)
    prepaid_notship_sheet = prepaid_notship_wb[prepaid_notship_wb.sheetnames[0]]
    # 抓還未銷貨的紀錄，若已銷貨，此資料會出現在會員日消費明細表
    prepaid_dict = dict()
    for row in prepaid_sheet.values:
        try:
            if row[2] is not None:
                datetime.datetime.strptime(row[2], date_format)
                # 沒有銷貨編號
                if row[12] is None:
                    # (row[1], row[3], row[4]): (店櫃名稱, 會員卡號, 銷貨編號(訂金))
                    prepaid_dict[(row[1], row[3], row[4])] = {
                        'date': row[2].replace('/', ''), # 銷貨日期(訂金)
                        'price': int(row[6]) + int(row[7]) # 訂金金額+尾款金額
                    }
        except ValueError:
            continue
    for row in prepaid_notship_sheet.values:
        # 假設銷貨編號全都是數字
        try:
            if row[3] is not None:
                int(row[3])
                store = ' '.join(row[2].split(' ')[1:]).strip()
                membercard = ' '.join(row[4].split(' ')[0])
                if (store, membercard, row[3]) in prepaid_dict:
                    date = prepaid_dict[(store, membercard, row[3])]
                    if date not in sales_dict:
                        sales_dict[date] = dict()
                    if store not in sales_dict[date]:
                        sales_dict[date][store] = {
                            'tickets': set(),
                            'sales': 0,
                            'cost': 0
                        }
                    sales_dict[date][store]['tickets'].add((prepaid_dict[(store, membercard, row[3])]['date'], store, membercard, row[3])) # 日期, 消費店櫃, 會員卡號, 銷貨編號
                    sales_dict[date][store]['sales'] += prepaid_dict[(store, membercard, row[3])]['price'] # 金額
                    sales_dict[date][store]['cost'] += product_dict[row[0]] * int(row[7]) # 成本 * 數量
        except ValueError:
            continue
    return sales_dict
'''
sales_dict = {
    '20210710' : {
        '士林店' : {
            'ticket' : 每張發票內容(集合),
            'sales' : 銷售金額,
            'cost' : 銷售商品之成本(用來算毛利(率))
        },
        '新店店' : {
            'ticket' : 發票內容(集合),
            'sales' : 銷售金額,
            'cost' : 銷售商品之成本(用來算毛利(率))
        },
        ...
    },
    '20210711' : {
        '士林店' : {
            'ticket' : 發票內容(集合),
            'sales' : 銷售金額,
            'cost' : 銷售商品之成本(用來算毛利(率))
        },
        '新店店' : {
            'ticket' : 發票內容(集合),
            'sales' : 銷售金額,
            'cost' : 銷售商品之成本(用來算毛利(率))
        },
        ...
    },
    ...
}
'''
# 計算Like for Like
def caculate_LFL(date, mode):
    # 初始化
    this_start, this_end, last_start, last_end = lineforline(date)
    dateList = []
    total_sales = 0
    total_margin = 0
    total_ticket = 0
    # 根據模式決定起始/結束日
    if mode == 'DAY':
        start_date = last_end
        end_date = last_end
    elif mode == 'WTD':        
        start_date, end_date =  weektodate(last_end)
    elif mode == 'Month':        
        start_date, end_date =  monthly(last_end)
    else:        
        start_date, end_date =  yeartodate(last_end)

    # 取得這期間的日期清單
    minus = end_date - start_date
    for date_count in range(0, minus.days+1):
        dateList.append(start_date + datetime.timedelta(days = date_count))

    # 取得這期間的sales, margin, ticket總和，若找不到某日則預設該日的值為0
    LFL_dict = dict()
    Store = Stores.objects.all()
    for stores in Store:
        LFL_dict[stores.name] = dict()
        LFL_dict[stores.name]['total_sales'] = 0
        LFL_dict[stores.name]['total_margin'] = 0
        LFL_dict[stores.name]['total_ticket'] = 0
        for date in dateList:
            if smt_report.objects.filter(date = date, stores = stores):
                Smt_report = smt_report.objects.get(date = date, stores = stores)
                store = Smt_report.stores.name
                LFL_dict[store]['total_sales'] += Smt_report.sales
                LFL_dict[store]['total_margin'] += (Smt_report.sales - Smt_report.margin)
                LFL_dict[store]['total_ticket'] += Smt_report.ticket_num
            else:
                LFL_dict[stores.name]['total_sales'] += 0
                LFL_dict[stores.name]['total_margin'] += 0
                LFL_dict[stores.name]['total_ticket'] += 0
    # 回傳值
    return LFL_dict
'''
LFL_dict = {
    '內湖店' : {
        total_sales : total_sales,
        total_margin : total_margin,
        total_ticket : total_ticket
    },
    ...
}
'''

# 在這裡一併處理budget, 與儲存進database的兩個步驟
# 決定報表內容(當日報表、WTD、YTD、月報表)
def group_sales(start_date, end_date, group_dict, sales_dict, mode):
    result_dict = dict()
    # budget_dict = dict()
    LFL_dict = dict()
    minus = end_date - start_date
    dateList = []
    for date_count in range(0, minus.days+1):
        dateList.append(start_date + datetime.timedelta(days = date_count))

    for date in dateList:
        if date in sales_dict: # 所需日期包含在上傳的資料中
        #for date in sales_dict:
            if date >= start_date and date <= end_date:
                for store in sales_dict[date]:
                    store = store.strip()
                    if group_dict[store]['primary'] not in result_dict:
                        result_dict[group_dict[store]['primary']] = dict()
                    if store not in result_dict[group_dict[store]['primary']]:
                        result_dict[group_dict[store]['primary']][store] = {
                            'tickets_num': 0,
                            'sales': 0,
                            'cost': 0
                        }
                    result_dict[group_dict[store]['primary']][store]['tickets_num'] += len(sales_dict[date][store]['tickets'])
                    result_dict[group_dict[store]['primary']][store]['sales'] += sales_dict[date][store]['sales']
                    result_dict[group_dict[store]['primary']][store]['cost'] += sales_dict[date][store]['cost']

        elif smt_report.objects.filter(date = date): # 若所需日不包含在上傳的資料中，從database找資料
            Smt_report = smt_report.objects.filter(date = date)
            for report in Smt_report:
                store = report.stores.name
                date = report.date

                if group_dict[store]['primary'] not in result_dict:
                    result_dict[group_dict[store]['primary']] = dict()
                if store not in result_dict[group_dict[store]['primary']]:
                    result_dict[group_dict[store]['primary']][store] = {
                        'tickets_num': 0,
                        'sales': 0,
                        'cost': 0
                    }
                result_dict[group_dict[store]['primary']][store]['tickets_num'] += Smt_report.ticket_num
                result_dict[group_dict[store]['primary']][store]['sales'] += Smt_report.sales
                result_dict[group_dict[store]['primary']][store]['cost'] += Smt_report.margin
        else:
            continue    
    LFL_dict = caculate_LFL(end_date, mode)

    return result_dict, LFL_dict


def save_report(start_date, end_date, group_dict, sales_dict, result_dict):
    # create/update smt_report data by date
    for date in sales_dict:
        if date >= start_date and date <= end_date:
            stores_list = Stores.objects.all().values_list('name')
            for primary in result_dict:
                for store in result_dict[primary]:
                    same_stores = Stores.objects.filter(name = store)
                    for stores in same_stores:
                        Budget = Stores_budget.objects.get(stores=stores)
                        if smt_report.objects.filter(date=date, stores=stores):
                            single_date = smt_report.objects.get(date=date, stores=stores)
                            single_date.date = date
                            single_date.sales = result_dict[group_dict[store]['primary']][store]['sales']
                            single_date.sales_budget = Budget.sales_budget
                            single_date.margin = result_dict[group_dict[store]['primary']][store]['cost']
                            single_date.margin_budget = Budget.margin_budget
                            single_date.ticket_num = result_dict[group_dict[store]['primary']][store]['tickets_num']
                            single_date.stores = stores
                            single_date.save()
                        else:
                            smt_report.objects.create(
                                date = date,
                                sales = result_dict[group_dict[store]['primary']][store]['sales'],
                                sales_budget = Budget.sales_budget,
                                margin = result_dict[group_dict[store]['primary']][store]['cost'],
                                margin_budget = Budget.margin_budget,
                                ticket_num = result_dict[group_dict[store]['primary']][store]['tickets_num'],
                                stores = stores,
                            )


'''
紀錄從start_date 到 end_date 的總發票數、總銷售數、總成本
result_dict = {
    '門市' : {
        '內湖店' : {
            '發票數' : 發票數
            '總銷售數' : 總銷售數
            '總成本' : 總成本
        },
        '士林店' : {
            '發票數' : 發票數
            '總銷售數' : 總銷售數
            '總成本' : 總成本
        },
        ...
    },
    '其它' : {
        'EC' : {
            '發票數' : 發票數
            '總銷售數' : 總銷售數
            '總成本' : 總成本
        },
        ...
    },
    ...
}
'''

# 輸出報表
def export_sales(result_dict, LFL_dict, template, output_file):
    # style
    gray_fill = PatternFill('solid', fgColor='D0CECE') # 灰色填滿
    bian = Side(style='thin', color='000000') # 設定邊框樣式
    border = Border(top=bian, bottom=bian, left=bian, right=bian)
    # 讀取模板
    wb_template = openpyxl.load_workbook(template)
    sales_sheet = wb_template[wb_template.sheetnames[0]]
    insert_row_count = 3
    total_rows = [] # 小計行數
    for primary in result_dict:
        start_row = insert_row_count
        stores_list = list()
        for store_name in result_dict[primary]:
            Sales = 0
            LFL_Sales = 0
            Cost = 0
            LFL_Margin = 0
            Tickets = 0
            LFL_Tickets = 0
            if store_name in stores_list:
                continue
            stores = Stores.objects.get(name = store_name)
            Secondary = Stores.objects.filter(subtype = stores.subtype)
            Budget = Stores_budget.objects.get(stores=stores)
            for secondary in Secondary:
                if secondary.name in result_dict[primary]:
                    Sales += result_dict[primary][secondary.name]['sales']
                    Cost += result_dict[primary][secondary.name]['cost']
                    Tickets += result_dict[primary][secondary.name]['tickets_num']
                    LFL_Sales += LFL_dict[secondary.name]['total_sales']                    
                    LFL_Margin += LFL_dict[secondary.name]['total_margin']                    
                    LFL_Tickets += LFL_dict[secondary.name]['total_ticket']
                    stores_list.append(secondary.name)
                elif secondary.name in LFL_dict:
                    LFL_Sales += LFL_dict[secondary.name]['total_sales']                    
                    LFL_Margin += LFL_dict[secondary.name]['total_margin']                    
                    LFL_Tickets += LFL_dict[secondary.name]['total_ticket']
                else:
                    continue

            sales_sheet.insert_rows(insert_row_count)
            sales_sheet['A'+str(insert_row_count)] = stores.subtype # Store
            sales_sheet['B'+str(insert_row_count)] = Sales # Sales $
            sales_sheet['C'+str(insert_row_count)] = Budget.sales_budget # Sales $ Budget
            sales_sheet['D'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('B'+str(insert_row_count), 'C'+str(insert_row_count)) # Sales$ ach%
            sales_sheet['E'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % (str(Sales), str(LFL_Sales)) # Sales$ LFL%
            sales_sheet['F'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('G'+str(insert_row_count), 'B'+str(insert_row_count)) # Margin%
            sales_sheet['G'+str(insert_row_count)] = Sales - Cost # Margin$
            sales_sheet['H'+str(insert_row_count)] = Budget.margin_budget # Margin$ Budget
            sales_sheet['I'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('G'+str(insert_row_count), 'H'+str(insert_row_count)) # Margin$ ach%
            sales_sheet['J'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % (str(Sales - Cost), str(LFL_Margin)) # Margin$ LFL%
            sales_sheet['K'+str(insert_row_count)] = Tickets # Tickets
            sales_sheet['L'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % (str(Tickets), str(LFL_Tickets)) # Tickets# LFL%
            insert_row_count += 1
        end_row = insert_row_count - 1
        for cell in sum(sales_sheet['A'+str(start_row):'L'+str(end_row)], ()):
            cell.border = border
        sales_sheet.insert_rows(insert_row_count)
        sales_sheet['A'+str(insert_row_count)] = primary + '小計'
        sales_sheet['D'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('B'+str(insert_row_count), 'C'+str(insert_row_count))
        sales_sheet['F'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('G'+str(insert_row_count), 'B'+str(insert_row_count))
        sales_sheet['I'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('G'+str(insert_row_count), 'H'+str(insert_row_count))
        for letter in list(string.ascii_uppercase)[0:12]:
            if letter in {'B', 'C', 'G', 'H', 'K'}:
                sales_sheet[letter+str(insert_row_count)] = '= SUM(%s:%s)' % (letter+str(start_row), letter+str(end_row))
            sales_sheet[letter+str(insert_row_count)].fill = gray_fill
            sales_sheet[letter+str(insert_row_count)].border = border
        total_rows.append(insert_row_count)
        insert_row_count += 1
    # 總計
    sales_sheet['D'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('B'+str(insert_row_count), 'C'+str(insert_row_count))
    sales_sheet['F'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('G'+str(insert_row_count), 'B'+str(insert_row_count))
    sales_sheet['I'+str(insert_row_count)] = '=ROUND((%s/%s) * 100, 1)' % ('G'+str(insert_row_count), 'H'+str(insert_row_count))
    for letter in ['B', 'C', 'G', 'H', 'K']:
        sales_sheet[letter+str(insert_row_count)] = '= SUM(%s)' % (', '.join([letter+str(c) for c in total_rows]))
    wb_template.save(output_file)

# 產生日銷售清單
def make_daily_sales_file(upload_file1, upload_file2, upload_file3):
    group_dict = dict()
    group_list = list()
    stores = Stores.objects.all()
    for token in stores:
        group_dict[token.name] = token.subtype
        if token.subtype not in group_list:
            group_list.append(token.subtype)
    group_list.append('其它')
    today = datetime.datetime.now()

    output_file = os.path.join(settings.BASE_DIR, 'output', '%s日銷售清單.xlsx' % today.strftime('%Y%m%d'))
    daily_sales_dict = daily_sales(upload_file1, upload_file2, upload_file3, group_dict)
    export_daily_sales(group_list, daily_sales_dict, 'excel_templates/daily_sales.xlsx', output_file) # 產生日銷售清單

# 產生Day/Month/WTD/YTD報表
def make_sales_file(upload_file1, upload_file2, upload_file3):
    group_dict = dict()
    group_list = list()
    product_dict = dict()

    sku = SKU.objects.all()
    stores = Stores.objects.all()
    for token in sku:
        product_dict[token.sku] = token.cost
    for token in stores:
        group_dict[token.name] = {
            'primary': token.category,
            'secondary': token.subtype
        }
    today = datetime.datetime.now()
    sales_dict = sales(product_dict, upload_file1, upload_file2, upload_file3)

    # DAY
    for date in sales_dict:
        
        start_date = date
        end_date = date
        result_dict, LFL_dict = group_sales(start_date, end_date, group_dict, sales_dict, 'DAY')
        # save_report(start_date, end_date, group_dict, sales_dict, result_dict)
        output_file = os.path.join(settings.BASE_DIR, 'output', '%sDAY銷售總表.xlsx' % start_date.strftime('%Y%m%d'))
        export_sales(result_dict, LFL_dict, 'excel_templates/sales.xlsx', output_file)
    # WTD
    for date in sales_dict:        
        start_date, end_date = weektodate(date)
        result_dict, LFL_dict = group_sales(start_date, end_date, group_dict, sales_dict, 'WTD')
        output_file = os.path.join(settings.BASE_DIR, 'output', '%s_%sWTD銷售總表.xlsx' % (start_date.strftime('%Y%m%d'), end_date.strftime('%Y%m%d')))
        # export_sales 前兩個輸入為sales_budget, margin_budget 待修正
        export_sales(result_dict, LFL_dict, 'excel_templates/sales.xlsx', output_file)

    # Month
    for date in sales_dict:
        if today.day == 1:
            start_date, end_date = monthly(date)
            result_dict, LFL_dict = group_sales(start_date, end_date, group_dict, sales_dict, 'Month')
            output_file = os.path.join(settings.BASE_DIR, 'output', '%s_%sMonth銷售總表.xlsx' % (start_date.strftime('%Y%m%d'), end_date.strftime('%Y%m%d')))
            # export_sales 前兩個輸入為sales_budget, margin_budget 待修正
            export_sales(result_dict, LFL_dict, 'excel_templates/sales.xlsx', output_file)
        else:
            continue
    # YTD
    for date in sales_dict:
        if date == datetime.datetime(2021, 6, 2) or date == datetime.datetime(2020, 5, 27):
            start_date, end_date = yeartodate(date)
            result_dict, LFL_dict = group_sales(start_date, end_date, group_dict, sales_dict, 'YTD')
            output_file = os.path.join(settings.BASE_DIR, 'output', '%s_%sYTD銷售總表.xlsx' % (start_date.strftime('%Y%m%d'), end_date.strftime('%Y%m%d')))
            export_sales(result_dict, LFL_dict, 'excel_templates/sales.xlsx', output_file)

#
def upload_smt(request):
    errors = list()
    if request.method == 'POST':
        upload_file1 = request.FILES['smt1']
        upload_file2 = request.FILES['smt2']
        upload_file3 = request.FILES['smt3']

        # 產生日銷售清單
        make_daily_sales_file(upload_file1, upload_file2, upload_file3)
        # 產生Day/Month/WTD/YTD報表
        # make_sales_file(upload_file1, upload_file2, upload_file3)
        #return redirect(reverse('smt_report:smt_report'))
        return render(request,'upload_smt.html', locals())
    return render(request,'upload_smt.html', locals())

# def smt_report_report(request):

