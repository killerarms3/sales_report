from django.shortcuts import render
from members.models import Members
from extra_table.models import CRM_tags
from django.conf import settings
from django.http import HttpResponse, Http404
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Side, Border
import os
import glob
# Create your views here.

def get_driver(executable_path):
    options = Options()
    options.add_argument('-headless')
    driver = webdriver.Firefox(executable_path=executable_path, options=options)
    return driver

# 去CRM抓特定區間內的會員數
def get_members(start=None, end=None):
    driver = get_driver(settings.PATH_GECKODRIVER)
    members_dict = {}
    try:
        driver.get('https://member.gsscloud.com/cas/login?service=https://www.videgree.com/23b0cab855374d7ebd4cb3eae90b0014/ExternalLogOn.mvc/SsoCasLoginCallback')
        driver.find_element_by_id('username').send_keys(settings.CRM_USERNAME)
        driver.find_element_by_id('password').send_keys(settings.CRM_PASSWORD)
        driver.find_element_by_name('submit').click()

        for tag in CRM_tags.objects.all():
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@id='LeftLabelQuickList']/div[@leftquicklabelid='Customer.All']"))).click()
            WebDriverWait(driver, 60).until(lambda driver: driver.execute_script("return jQuery.active == 0"))
            # time.sleep(5)
            # total
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='footer']/button[@class='btn more-btn']"))).click()
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='title']/button[@class='title-btn add']"))).click()
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='LabelItems']/label[text()='%s']" % (tag.tag)))).click()
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'LabelConditionApply'))).click()
            WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, 'CreateDate'))).clear()
            if (start is not None) and (end is not None):
                WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, 'CreateDate'))).send_keys(start.strftime('%Y/%m/%d') + ' - ' + end.strftime('%Y/%m/%d'))
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='range_inputs']/button[@class='applyBtn hide btn btn-sm btn-success']"))).click()
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'QueryBtn'))).click()
            WebDriverWait(driver, 60).until(lambda driver: driver.execute_script("return jQuery.active == 0"))
            result_text =  WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div[@id='pager']/span[@class='k-pager-info k-label']"))).text
            try:
                total_member = int(re.search(r'共\s([0-9]+)\s筆', result_text).group(1))
            except AttributeError:
                total_member = 0
            if total_member != 0:
                # 消費次數大於0的會員
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='footer']/button[@class='btn more-btn']"))).click()
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.filter-group.order.hide-content'))).find_element_by_css_selector('button.filter-collapse-btn').click()
                WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, 'ConsumerTimes'))).send_keys('0')
                WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'QueryBtn'))).click()
                WebDriverWait(driver, 60).until(lambda driver: driver.execute_script("return jQuery.active == 0"))
                result_text =  WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div[@id='pager']/span[@class='k-pager-info k-label']"))).text
                try:
                    has_order = int(re.search(r'共\s([0-9]+)\s筆', result_text).group(1))
                except AttributeError:
                    has_order = 0
                if has_order != 0:
                    # 有消費但沒手機的會員
                    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='footer']/button[@class='btn more-btn']"))).click()
                    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.filter-group.customercontent.hide-content'))).find_element_by_css_selector('button.filter-collapse-btn').click()
                    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'select2-TelecomNumOption-container'))).click()
                    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//ul[@class='select2-results__options']/li[3]"))).click()
                    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='check-condition']/div[@class='checkbox-item']/label[@for='ReverseQueryTelecomNum']"))).click()
                    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'QueryBtn'))).click()
                    WebDriverWait(driver, 60).until(lambda driver: driver.execute_script("return jQuery.active == 0"))
                    result_text =  WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div[@id='pager']/span[@class='k-pager-info k-label']"))).text
                    try:
                        nophone_has_order = int(re.search(r'共\s([0-9]+)\s筆', result_text).group(1))
                    except AttributeError:
                        nophone_has_order = 0
            else:
                has_order = 0
                nophone_has_order = 0
            members_dict[tag.tag] = {
                'total_member': total_member,
                'valid_member': has_order - nophone_has_order
            }
    except TimeoutException:
        # CRM帳密要到期需要更換
        driver.quit()
    finally:
        driver.quit()
    return members_dict


def save_members():
    today = datetime.datetime.now()
    last_week_start = today - datetime.timedelta(days=today.weekday() + 7)
    last_week_end = today - datetime.timedelta(days=today.weekday() + 1)
    total_member = get_members(start=datetime.datetime(2020,12,22), end=last_week_end)
    week_member = get_members(start=last_week_start, end=last_week_end)
    for tag in CRM_tags.objects.all():
        date = (last_week_end - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        if not Members.objects.filter(date=date, label=tag.tag):
            member = Members()
        else:
            member = Members.objects.filter(date=date, label=tag.tag)[0]
        member.date = date
        member.new = week_member[tag.tag]['total_member']
        member.total = total_member[tag.tag]['total_member']
        member.eff_new = week_member[tag.tag]['valid_member']
        member.eff_total = total_member[tag.tag]['valid_member']
        member.label = tag
        member.save()

# Group結果
def group_result(date):
    result_dict = dict()
    for tag in CRM_tags.objects.all():
        members = Members.objects.filter(date=date.strftime('%Y-%m-%d'), label=tag.tag)
        if tag.category not in result_dict:
            result_dict[tag.category] = dict()
        if tag.subtype not in result_dict[tag.category]:
            result_dict[tag.category][tag.subtype] = {
                'new': 0,
                'total': 0,
                'eff_new': 0,
                'eff_total': 0
            }
        if members:
            member = members[0]
            result_dict[tag.category][tag.subtype]['new'] += member.new
            result_dict[tag.category][tag.subtype]['total'] += member.total
            result_dict[tag.category][tag.subtype]['eff_new'] += member.eff_new
            result_dict[tag.category][tag.subtype]['eff_total'] += member.eff_total
    return result_dict

def export_members(last_1_week_date):
    # last_1_week_date: datetime (monday)
    template = os.path.join(settings.BASE_DIR, 'excel_templates', 'members.xlsx')
    output_file = os.path.join(settings.BASE_DIR, 'output', 'Members%s.xlsx' % last_1_week_date.strftime('%Y%m%d'))

    # 取得會員數
    # 由資料庫查詢上週、上上週與上上上週會員數
    last_2_week_date = last_1_week_date - datetime.timedelta(days=7)
    last_3_week_date = last_1_week_date - datetime.timedelta(days=14)
    last_1_week_result = group_result(last_1_week_date)
    last_2_week_result = group_result(last_2_week_date)
    last_3_week_result = group_result(last_3_week_date)

    # 讀取模板
    wb_template = openpyxl.load_workbook(template)
    member_sheet = wb_template['會員']
    valid_member_sheet = wb_template['有效會員']
    insert_row_count = 3 #由第三行前開始insert
    total_rows = [] # 小計行數
    gray_fill = PatternFill('solid', fgColor='D0CECE') # 灰色填滿
    bian = Side(style='thin', color='000000') # 設定邊框樣式
    border = Border(top=bian, bottom=bian, left=bian, right=bian)
    for primary in last_1_week_result:
        start_row = insert_row_count
        for secondary in last_1_week_result[primary]:
            # 插入空行
            member_sheet.insert_rows(insert_row_count)
            valid_member_sheet.insert_rows(insert_row_count)
            # 填入會員數
            member_sheet['A'+str(insert_row_count)] = secondary
            member_sheet['B'+str(insert_row_count)] = last_1_week_result[primary][secondary]['new']
            member_sheet['C'+str(insert_row_count)] = last_2_week_result[primary][secondary]['new'] # 上上週會員
            member_sheet['D'+str(insert_row_count)] = last_3_week_result[primary][secondary]['new'] # 上上上週會員
            member_sheet['E'+str(insert_row_count)] = last_1_week_result[primary][secondary]['total']
            valid_member_sheet['A'+str(insert_row_count)] = secondary
            valid_member_sheet['B'+str(insert_row_count)] = last_1_week_result[primary][secondary]['eff_new']
            valid_member_sheet['C'+str(insert_row_count)] = last_2_week_result[primary][secondary]['eff_new']
            valid_member_sheet['D'+str(insert_row_count)] = last_3_week_result[primary][secondary]['eff_new']
            valid_member_sheet['E'+str(insert_row_count)] = last_1_week_result[primary][secondary]['eff_total']
            insert_row_count += 1
        end_row = insert_row_count - 1
        for cell in sum(member_sheet['A'+str(start_row):'E'+str(end_row)], ()):
            cell.border = border
            valid_member_sheet[cell.coordinate].border = border
        member_sheet.insert_rows(insert_row_count)
        valid_member_sheet.insert_rows(insert_row_count)

        for letter in ['A', 'B', 'C', 'D', 'E']:
            if letter == 'A':
                member_sheet['A'+str(insert_row_count)] = primary + '小計'
                valid_member_sheet['A'+str(insert_row_count)] = primary + '小計'
            else:
                member_sheet[letter+str(insert_row_count)] = '= SUM(%s:%s)' % (letter+str(start_row), letter+str(end_row))
                valid_member_sheet[letter+str(insert_row_count)] = '= SUM(%s:%s)' % (letter+str(start_row), letter+str(end_row))
            member_sheet[letter+str(insert_row_count)].fill = gray_fill
            valid_member_sheet[letter+str(insert_row_count)].fill = gray_fill
            member_sheet[letter+str(insert_row_count)].border = border
            valid_member_sheet[letter+str(insert_row_count)].border = border
        total_rows.append(insert_row_count)
        insert_row_count += 1
    for letter in ['B', 'C', 'D', 'E']:
        member_sheet[letter+str(insert_row_count)] = '= SUM(%s)' % (', '.join([letter+str(c) for c in total_rows]))
        valid_member_sheet[letter+str(insert_row_count)] = '= SUM(%s)' % (', '.join([letter+str(c) for c in total_rows]))
    wb_template.save(output_file)

def members_report(request):
    reports = dict()
    for file_path in glob.glob(os.path.join(settings.BASE_DIR, 'output', 'Members*.xlsx')):
        filename = os.path.basename(file_path)
        date = re.findall(r'\d+', filename)[0]
        reports[date] = filename
    return render(request,'members_report.html', locals())


def download(request, filename):
    file_path = os.path.join(settings.BASE_DIR, 'output', filename)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404
